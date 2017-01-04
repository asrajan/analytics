"""
module : write_table

Exports a table as a readable excel sheet in xlsx format
"""
from .table import ATable
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Color, colors
from openpyxl.utils import get_column_letter
import logging


class AWriteTable(object):
    """ Writes the table out in xlsx format.

    This class implements the translation of a table into
    xlsx format. It assumes that the table is valid.

    Attributes:
        table : A reference to the ATable object
        file  : Either name of the file to be writen out or a file object
    """
    def __init__(self, table, file):
        self._file = file
        self._table = table
        # Additional validation - Maybe superfluous
        self._table._validate()
        # Cache Workbook
        _wb = None

    def execute(self):
        """ Engine method that transforms the table to xlsx format 
        Raises:
            PermissionError
        """
        # Open a work book with optimized_write
        self._wb = Workbook()
        ws = self._wb.active
        #hfill = PatternFill(fill_type='solid',
        #        fgColor = Color('00777777'),
        #        bgColor = Color('00777777'))
        hfont = Font(color=colors.BLACK, bold=True)

        # The Column Heading Row
        for idx,value in enumerate(self._table.columns):
            cell = ws[get_column_letter(idx+1) + '1']
            cell.font = hfont
            #cell.fill = hfill
            #cell.border = hborder
            cell.value = value
        # Pushing data
        for row in self._table._data:
            ws.append(row)
       
        self._widen_cols()
        self._colorize_rows()
        self._data_type_cols()
        
        try:
            self._wb.save(self._file)
        except PermissionError:
            logging.info('''Failed to create an Excel file. Probably opened.''')
            raise
    
    

    def _translate_type(self, cell_obj, type_string):
        ''' Translates the type code to a valid cell type '''
        if type_string == ATable.STRING:
            if cell_obj.value: 
                if cell_obj.value.isnumeric():
                    cell_obj.value = '=("' + cell_obj.value +'")'
            return
        if type_string == ATable.NUMBER:
            #cell_obj.set_explicit_value(cell_obj.value, 'n')
            return
        if type_string == ATable.BINARY:
            raise RuntimeError('Unhandled format Binary')
            return
        if type_string == ATable.DATETIME:
            #cell_obj._bind_value(cell_obj.value)
            return
        if type_string == ATable.ROWID:
            raise RuntimeError('Unhandled format ROWID')
            return
        raise RuntimeError('Unknown format : ' + type_string)
    
    def _as_text(self, value):
        ''' Converts to string. Treats None as empty string '''
        if value == None:
            return ""
        else:
            return str(value)
            
    def _widen_cols(self):
        ''' Sets the columns to a width size that is easily 
        readable 
        
        Entries in different fields of a column will require
        different column budget so that it can be easily readable.
        This is based on an an estimate only.
        
        This must be run before _translate_type is called.
        '''
        ws = self._wb.active
        # Initialize the column width to be 10
        min_width = 10
        # Font width fudge factor
        ff = 1.25
        # Iterate over columns
        for col_cells in ws.columns:
            length = max(
                max(
                len(self._as_text(cell.value))*ff for cell in col_cells),
                min_width)
            ws.column_dimensions[col_cells[0].column].width = length
        
    def _colorize_rows(self):
        ''' Colors the rows of the sheet.
        
        Colors are specified in XXRRGGBB format. The Hex colors can be 
        obtained from - http://www.color-hex.com/color-names.html 
        '''
        ws = self._wb.active
        # Define style constants
        sky_blue = Color('00CDE2ED')
        deep_blue = Color('0087CEEB')
        white = Color('FFFFFFFF')
        dark_sky_blue = Color('008CBED6')
        odd_row_fill = PatternFill(fill_type='solid', 
                fgColor = sky_blue,
                bgColor = sky_blue)
        even_row_fill =  PatternFill(fill_type='solid', 
                fgColor = white,
                bgColor = white)
        side = Side(border_style='thin', color=dark_sky_blue)
        hborder = Border(left=side, right=side, top=side, bottom=side)
        # Title Row should look different
        title_row_fill = PatternFill(fill_type='solid', 
                fgColor = deep_blue,
                bgColor = deep_blue)
        for row in ws.iter_rows(min_row=1,max_row=1):
            for cell in row:
                cell.fill = title_row_fill
                cell.border = hborder
                
               
        cell_fills = [even_row_fill, odd_row_fill]
        ridx = 0
        for row in ws.iter_rows(min_row=2, max_col=len(self._table.columns),
                max_row = len(self._table.data)+1):
            for cell in row:
                cell.fill = cell_fills[ridx%2]
                cell.border = hborder
            ridx = ridx+1
    
    def _data_type_cols(self):
        ''' Set the data type on the columns if it is known in the table
        
        Each column has a data type associated with the entries. These 
        data types are useful to enable the spread sheet tools to interpret
        the data correctly.
        '''
        ws = self._wb.active
        if not self._table.types:
            return
        for row in ws.iter_rows(min_row=1, max_col=len(self._table.columns),
            max_row = len(self._table.data)+1):
            col_idx = 0
            for cell in row:
                self._translate_type(cell, self._table.types[col_idx])
                col_idx = col_idx+1





