"""
module : artworkasset.py

Extracts information from an excel sheet at envogue that stores the 
cost information for each art work.
"""
from openpyxl import load_workbook
from datamodel.error import AError
from nltk.metrics import edit_distance

class ArtworkAsset(object):
    """ This is very specific to envogue.

    The Finance department has so far maintained some asset management 
    information for artwork. This is the only place to correlate the 
    pricing of the artwork with digital files.

    This is really a data base and we must optimize for indices. So far
    this is very small so we reall do not need to do that.
    
    """
    SHEET_LOCATION = "K:\\Artworks and Samples Listing\\Artwork Asset Listing.xlsx"
    ACTIVE_SHEET = "Artwork"
    _DATABASE = dict()
    def __init__(self):
        if ArtworkAsset._DATABASE:
            return
        wb = load_workbook(
                ArtworkAsset.SHEET_LOCATION, read_only=True)
        ws = wb[ArtworkAsset.ACTIVE_SHEET]
        all_cells = ws['A4':('G'+str(ws.max_row))]
        for row in all_cells:
            label = row[0].value
            original_name = row[1].value 
            if not original_name:
                continue
            name = row[2].value 
            studio = row[3].value 
            invoice_date = row[4].value 
            invoice_num = row[5].value 
            cost = row[6].value 
            db_entry = dict();
            db_entry["name"] = name
            db_entry["studio"] = studio
            db_entry["invoice_date"] = invoice_date
            db_entry["cost"] = cost
            db_entry["invoice_num"] = invoice_num
            if not original_name in ArtworkAsset._DATABASE:
                ArtworkAsset._DATABASE[original_name] = db_entry
            else:
                raise AError("Duplicate entries found - " + original_name)
            ArtworkAsset._DATABASE[original_name] = db_entry

    def cost(self, name, studio):
        """ Returns the cost associated with an artwork name

        Uses Fuzzy search to locate the information.
        """
        key = self.search(name, studio, 5)
        if key:
            return ArtworkAsset._DATABASE[key]["cost"]
        return None

    def search(self, name, studio=None, max_tries=5):
        """ Fuzzy match algorithm to pick the most likely candidate
        
        Searches through the keys in a dictionary and locates the most likely
        candidate.

        First the keys are normalized by eliminating white spaces
        The value to be searched is normalized by eliminating white spaces
        Both the keys and the values are made lower case to eliminate case 
        sensitivity.

        Length of all the keys is truncated to be the same length as the name
        to minimize suffix related issues.

        """
        # Normalize the name to be searched
        name = (''.join(name.split())).lower()
        names_orig = [n for n in ArtworkAsset._DATABASE.keys()]
        # Normalize the keys
        names = [''.join(n.lower().split()) for n in ArtworkAsset._DATABASE.keys()]
        names = [n if len(n)<len(name) else n[:len(name)-1] for n in names]
        # Compute the edit distances
        distance_index = list(
                enumerate([edit_distance(n, name) for n in names]))
        sorted_distance_index = sorted(distance_index,key=lambda distance: distance[1])
        if studio:
            for i in range(0,max_tries-1):
                if ArtworkAsset._DATABASE[names_orig[sorted_distance_index[i][0]]]["studio"] == studio:
                    return names_orig[sorted_distance_index[i][0]]
            return None
        return names_orig[sorted_distance_index[i][0]]






